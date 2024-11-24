import xgboost as xgb
from sklearn.datasets import load_breast_cancer
from sklearn.model_selection import train_test_split, GridSearchCV, RandomizedSearchCV
from sklearn.metrics import accuracy_score, roc_auc_score, classification_report, roc_curve, precision_score, recall_score, f1_score
import pandas as pd
import numpy as np
from xgboost import plot_importance
import matplotlib.pyplot as plt
import openpyxl
from sklearn.metrics import confusion_matrix, ConfusionMatrixDisplay
import shap

# Load dataset
df = pd.read_excel(r'C:\Users\Administrator\Desktop\shuju\T80 and T90.xlsx', sheet_name='fenlei')
df = df.dropna()
X = df.iloc[:, 0:-1]
y = df.iloc[:, -1]

# Split the dataset into training and testing sets
X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2, random_state=42)

# Set hyperparameters to be tuned
param_grid = {
    'max_depth': (3, 20),
    'min_child_weight': (1, 20),
    'gamma': (0, 0.5),
    'subsample': (0.5, 0.9),
    'colsample_bytree': (0.5, 0.9),
    'reg_alpha': (1e-5, 1),
    'reg_lambda': (1e-5, 100),
    'learning_rate': (0.001, 1),
    'n_estimators': (100, 1000)
}

# Create XGBClassifier object
model = xgb.XGBClassifier(objective='binary:logistic', early_stopping_rounds=100)

# Use GridSearchCV for hyperparameter tuning
grid_search = GridSearchCV(estimator=model, param_grid=param_grid, scoring='f1', cv=10, n_jobs=-1, verbose=1)
grid_search.fit(X_train, y_train, eval_set=[(X_test, y_test)], verbose=False)

# Output the best parameter combination
print("Best parameters: ", grid_search.best_params_)
print("Best score: ", grid_search.best_score_)

# Predict on the test set
y_pred = grid_search.predict(X_test)

# Validation
param = {
    'max_depth': 3,
    'min_child_weight': 1,
    'gamma': 0,
    'subsample': 0.5,
    'colsample_bytree': 0.9,
    'reg_alpha': 1,
    'reg_lambda': 100,
    'learning_rate': 0.001
}
model = xgb.XGBClassifier(objective='multi:softmax', num_class=2, n_estimators=100, early_stopping_rounds=10)
# Apply hyperparameters
model = model.set_params(**param)
bst = model.fit(X_train, y_train, eval_set=[(X_test, y_test)], verbose=False)
y_pred = model.predict(X_test)

# Visualize feature importance
plot_importance(bst, max_num_features=40, importance_type='weight', title='Feature Importance')
plt.show()
feature_importance = model.get_booster().get_score(importance_type='weight')

# Sort feature importance
sorted_feature_importance = sorted(feature_importance.items(), key=lambda x: x[1], reverse=True)

# Print sorted feature importance
print("Sorted Feature Importance:")
for feature, importance in sorted_feature_importance:
    print(f"Feature {feature}: {importance}")

# Create a new Excel workbook
workbook = openpyxl.Workbook()
worksheet = workbook.active

# Write sorted feature importance to Excel
worksheet['A1'] = 'Feature'
worksheet['B1'] = 'Importance'
for row, (feature, importance) in enumerate(sorted_feature_importance, start=2):
    worksheet.cell(row=row, column=1, value=f"Feature {feature}")
    worksheet.cell(row=row, column=2, value=importance)

# Save the Excel file
workbook.save('feature_importance.xlsx')

# Calculate accuracy
accuracy = accuracy_score(y_test, y_pred)
print(f"Accuracy: {accuracy}")

# Calculate confusion matrix
cm = confusion_matrix(y_test, y_pred, normalize='true')
print("Confusion Matrix:\n", cm)

# Plot confusion matrix
disp = ConfusionMatrixDisplay(confusion_matrix=cm)
disp.plot(cmap=plt.cm.Blues)
plt.show()

y_prob = model.predict_proba(X_test)[:, 1]  # Get the probability of the positive class
fpr, tpr, thresholds = roc_curve(y_test, y_prob)
plt.figure()
plt.plot(fpr, tpr, label='ROC Curve')
plt.plot([0, 1], [0, 1], 'k--')
plt.xlabel('False Positive Rate')
plt.ylabel('True Positive Rate')
plt.title('ROC Curve')
plt.legend()
plt.show()
accuracy = accuracy_score(y_test, y_pred)
precision = precision_score(y_test, y_pred)
recall = recall_score(y_test, y_pred)
f1 = f1_score(y_test, y_pred)
auc = roc_auc_score(y_test, y_prob)
print(f"Accuracy: {accuracy:.3f}, Precision: {precision:.3f}, Recall: {recall:.3f}, F1-score: {f1:.3f}, AUC: {auc:.3f}")

# Use SHAP to calculate feature importance
explainer = shap.TreeExplainer(model)
shap_values = explainer.shap_values(X_test, check_additivity=False)
class_names = ['T90<1000h', 'T90>1000h']
# Summarize feature importance (mean)
shap.summary_plot(shap_values, X_test, plot_type="bar", max_display=19, class_names=class_names)

# Plot feature importance (all data instances)
shap.summary_plot(shap_values[0], X_test, plot_type="dot", max_display=19)

# Explain a single prediction instance
idx = 10  # Select the instance index to explain
shap.force_plot(explainer.expected_value[1], shap_values[1][idx,:], X_test.iloc[idx], matplotlib=True, link='identity')

shap.dependence_plot('Perovskite_bandgap', shap_values[1], X_train, interaction_index=None, show=True)